# a_xlsx.py

### 파이썬의 데이터를 활용하여 엑셀 파일 생성 ###
#! pandas 라이브러리 사용

import pandas as pd

data = {
    'Name': ['이승아', '이도경', '김명진'],
    'Age': [29, 31, 35],
    'City': ['부산', '대전', '서울']
}

# 데이터 프레임 생성
# : pandas 라이브러리의 DataFrame() 함수 사용
df = pd.DataFrame(data)

# 데이터 프레임 > 엑셀 파일로 저장
# cf) index=False
#       : pandas가 엑셀 파일로 데이터를 저장할 때, 인덱스를 파일에 포함 X
# : 데이터프레임변수명.to_excel('엑셀파일명.xlsx', index=False)
df.to_excel('toExcel0207.xlsx', index=False)

#? cf) pandas 라이브러리 실행 시 사용하고 있는 엑셀 파일이 열려 있는 경우
# : permission Error가 발생

### openpyxl - 파이썬에서 엑셀 데이터 다루기 ###
# : loadworkbook 모듈 활용
#   - 지정된 경로에 있는 엑셀 파일 호출 / 파이썬에서 다룰 수 있는 객체 형식으로 변경

from openpyxl import load_workbook

# 엑셀 파일 불러오기
# : load_workbook('파일명.확장자')
wb = load_workbook('toExcel0207.xlsx')
print(wb) # 엑셀 파일이 담긴 메모리 주소값이 출력

# 기존에 있는 첫 번째 시트 선택
sheet = wb.active
print(sheet) # <Worksheet "Sheet1">

# cf) 특정 이름의 시트 선택
# selected_sheet = wb['시트명'] # wb['Sheet1']

# cf) 특정 위치에 있는 시트에 접근
# : 위치 - 전체 시트가 리스트로 저장 (인덱스 번호)
# selected_sheet = wb.worksheets['위치번호(숫자)']
# wb.workshhets[0] (Sheet1)

### openpyxl을 활용한 수정 ###
from openpyxl.styles import Font, Color, Alignment, PatternFill
# Font(글꼴), Color(색상), Alignment(배치, 정렬), PatternFill(셀 배경 채우기)

# 첫 번째 행(제목 행)의 폰트와 정렬 변경
# : 첫 번째 행에 있는 모든 셀에 대한 반복
for cell in sheet['1:1']:
    # 각 셀의 스타일 지정
    # cell변수명.font = Font모듈(속성 지정)
    # > bold(두께), color(색상-16진수)
    cell.font = Font(bold=True, color='FFFFFF') # 흰색
    cell.alignment = Alignment(horizontal='center') # 수직으로 가운데 정렬
    
    # 배경을 단색 지정
    # : PatternFill은 그라데이션 색상을 지정
    # > start_color와 end_color를 동일하게 지정하는 경우 단색 지정 가능
    cell.fill = PatternFill(start_color='0000ff', end_color='0000ff', fill_type='solid')
    
# cf) 16진수: 0 ~ 9와 a ~ f까지의 문자로 수를 표현
#       - 주로 색상 코드에 사용

# 각 열의 너비를 조정
# : 시트명.column_dimensions['열기호'].width = 너비
sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 30
sheet.column_dimensions['C'].width = 10

# 행의 높이 조정
# : 시트명.row_dimensions[행번호].height = 높이
sheet.row_dimensions[1].height = 20

# 동일한(존재하는) 파일을 업데이트
# : save('파일명.확장자')
wb.save('toExcel0207_modified.xlsx')

wb.close() # 엑셀 파일 닫기